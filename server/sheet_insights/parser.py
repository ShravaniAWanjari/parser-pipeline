import re
from pathlib import Path
import openpyxl
from markitdown import MarkItDown

def get_sheet_names(file_path):
    """Extract all sheet names from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        names = workbook.sheetnames
        workbook.close()
        print(f"📋 Found {len(names)} sheets: {names}")
        return names
    except Exception as e:
        print(f"❌ Failed to load sheet names: {e}")
        return []

def format_metadata(sheet):
    """Extract and format metadata from the top few rows of the Excel sheet"""
    metadata = []
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        line = ' '.join([str(cell).strip() for cell in row if cell])
        if line:
            metadata.append(line)
    return metadata

def is_empty_row(row):
    """Check if a row is essentially empty (all None, empty strings, or whitespace)"""
    return all(cell is None or str(cell).strip() == '' for cell in row)

def has_meaningful_content(row):
    """Check if row has meaningful content (not just formulas or minimal data)"""
    non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip()]
    
    if len(non_empty_cells) < 2:
        return False
    
    formula_only = all(str(cell).startswith('=') for cell in non_empty_cells)
    if formula_only and len(non_empty_cells) < 3:
        return False
    
    return True

def find_data_boundaries(sheet, start_row=6):
    """Find the actual data boundaries to avoid processing too many empty rows"""
    rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
    
    last_meaningful_row = -1
    for i, row in enumerate(rows):
        if has_meaningful_content(row):
            last_meaningful_row = i
    
    if last_meaningful_row >= 0:
        return rows[:last_meaningful_row + 4] 
    else:
        return rows[:10]

def extract_table(sheet):
    """Extract and format the core table starting from a specific row"""
    start_row = 6  # Adjust based on where your actual data starts
    
    rows = find_data_boundaries(sheet, start_row)
    
    if not rows:
        return []
    
    non_empty_rows = [row for row in rows if not is_empty_row(row)]
    if not non_empty_rows:
        return []
    
    max_cols = max(len(row) for row in non_empty_rows)
    
    # Process rows and remove consecutive empty rows (keep max 3-4)
    processed_rows = []
    consecutive_empty = 0
    max_consecutive_empty = 3
    
    for row in rows:
        # Pad row to match max columns
        padded_row = list(row) + [None] * (max_cols - len(row))  # Use None instead of ''
        
        if is_empty_row(padded_row):
            consecutive_empty += 1
            if consecutive_empty <= max_consecutive_empty:
                processed_rows.append(padded_row)
        else:
            consecutive_empty = 0
            processed_rows.append(padded_row)
    
    # Convert to markdown
    markdown_lines = []
    for i, row in enumerate(processed_rows):
        # Clean cell content
        cells = []
        for cell in row:
            if cell is None or cell == '':
                cells.append('null')  # Use 'null' string for empty cells
            else:
                # Clean the cell content
                cell_str = str(cell).replace('\n', ' ').strip()
                # Handle Excel errors
                if cell_str.startswith('#') or cell_str == '#DIV/0!':
                    cells.append('null')  # Convert Excel errors to null
                # Truncate very long formula strings
                elif cell_str.startswith('=') and len(cell_str) > 50:
                    cells.append('null')  # Convert long formulas to null
                else:
                    cells.append(cell_str)
        
        line = '| ' + ' | '.join(cells) + ' |'
        markdown_lines.append(line)
        
        # Add header separator after first row if it seems like a header
        if i == 0 and has_meaningful_content(row):
            markdown_lines.append('| ' + ' | '.join(['---'] * len(cells)) + ' |')
    
    return markdown_lines

def clean_markdown_post_process(markdown_path, max_empty_lines=3):
    """Post-process markdown file to remove excessive empty table rows"""
    try:
        with open(markdown_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        cleaned_lines = []
        consecutive_empty_table_rows = 0
        
        for line in lines:
            stripped = line.strip()
            if stripped.startswith('|') and stripped.endswith('|'):
                content = stripped[1:-1].split('|')
                if all(cell.strip() == '' for cell in content):
                    consecutive_empty_table_rows += 1
                    if consecutive_empty_table_rows <= max_empty_lines:
                        cleaned_lines.append(line)
                else:
                    consecutive_empty_table_rows = 0
                    cleaned_lines.append(line)
            else:
                consecutive_empty_table_rows = 0
                cleaned_lines.append(line)
        
        with open(markdown_path, 'w', encoding='utf-8') as f:
            f.writelines(cleaned_lines)
        
        print(f"🧹 Cleaned excessive empty rows from {markdown_path.name}")
        
    except Exception as e:
        print(f"⚠️ Failed to clean {markdown_path}: {e}")

def extract_markdown(file_path, output_dir, sheets_to_process=None, skip_first_sheet=True, max_empty_rows=3):
    all_sheet_names = get_sheet_names(file_path)
    if not all_sheet_names:
        return [], {}

    if sheets_to_process:
        target_sheets = sheets_to_process
    else:
        target_sheets = all_sheet_names[1:] if skip_first_sheet else all_sheet_names

    print(f"📋 Processing {len(target_sheets)} sheet(s): {target_sheets}")

    markdown_paths = []
    name_mapping = {}

    md = MarkItDown(enable_plugins=False)
    result = md.convert(str(file_path))
    text = result.text_content

    sheet_heading_pattern = re.compile(r"^(#+)\s*(.+)$")
    current_sheet = None
    current_lines = []
    for line in text.splitlines():
        match = sheet_heading_pattern.match(line)
        if match:
            if current_sheet and current_lines and current_sheet in target_sheets:
                clean_sheet_name = re.sub(r'[^\w\-_]', '_', current_sheet.strip())
                clean_sheet_name = re.sub(r'_+', '_', clean_sheet_name).strip('_')
                markdown_path = output_dir / f"{clean_sheet_name}.md"
                with open(markdown_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(current_lines))
                markdown_paths.append(markdown_path)
                name_mapping[clean_sheet_name] = current_sheet
            current_sheet = match.group(2)
            current_lines = [line]
        else:
            if current_lines is not None:
                current_lines.append(line)
    if current_sheet and current_lines and current_sheet in target_sheets:
        clean_sheet_name = re.sub(r'[^\w\-_]', '_', current_sheet.strip())
        clean_sheet_name = re.sub(r'_+', '_', clean_sheet_name).strip('_')
        markdown_path = output_dir / f"{clean_sheet_name}.md"
        with open(markdown_path, "w", encoding="utf-8") as f:
            f.write("\n".join(current_lines))
        markdown_paths.append(markdown_path)
        name_mapping[clean_sheet_name] = current_sheet

    print(f"✅ Saved {len(markdown_paths)} markdown files using MarkItDown")
    return markdown_paths, name_mapping