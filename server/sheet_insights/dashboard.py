import openpyxl
import json
from pathlib import Path
from collections import defaultdict
import os
from openai import AzureOpenAI
from datetime import datetime


def generate_dashboard_json_with_llm(raw_supplier_kpi_data):
    """Send extracted KPI data to Azure OpenAI to return structured dashboard JSON."""
    today = datetime.today().strftime("%Y-%m-%d")

    system_prompt = f"""
You are an expert assistant that converts messy KPI data from multiple suppliers into a standardized structured JSON dashboard.
Use this template:
{{
  "generatedOn": "2025-07-30",
  "kpiMetadata": {{
    "unitDescriptions": {{
      "accidents": "Number of safety incidents reported",
      "productionLossHrs": "Production hours lost due to supplier-caused material shortage",
      "okDeliveryPercent": "Percentage of OK deliveries based on ACMA standards",
      "trips": "Number of shipment trips completed per month",
      "quantityShipped": "Number of parts shipped by the supplier",
      "partsPerTrip": "Efficiency metric showing avg. parts shipped per trip",
      "vehicleTAT": "Average vehicle turnaround time at the plant (in hours)",
      "machineDowntimeHrs": "Machine breakdown time (in hours)",
      "machineBreakdowns": "Number of machine breakdowns",
      "onTimeDeliveryRate": "Percentage of deliveries that arrived on or before schedule",
      "capaClosureDays": "Average number of days to close CAPA (Corrective Actions)",
      "costPerPart": "Logistics cost per part shipped",
      "supplierContributionPercent": "Supplier’s contribution to total parts supply"
    }}
  }},
  "suppliers": [
    {{
      "name": "Ankita Autocoters",
      "location": "Ranjangaon",
      "buyer": "Mr. Mahesh Shirsagar",
      "spoc": "Mr. Rohan Dhamale",
      "kpis": [
        {{
          "id": 1,
          "name": "Safety - Accident Data",
          "alias": "accidents",
          "unit": "nos",
          "monthly": [0, 0, 0, 0, null, null, null, null, null, null, null, null],
          "average": 0,
          "responsible": "Mr. Rohan",
          "remarks": ""
        }},
        {{
          "id": 2,
          "name": "Production Loss due to Material Shortage",
          "alias": "productionLossHrs",
          "unit": "hrs",
          "monthly": [0, 0, 0, 0, null, null, null, null, null, null, null, null],
          "average": 0,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 3,
          "name": "OK Delivery Cycles (ACMA Standard)",
          "alias": "okDeliveryPercent",
          "unit": "%",
          "monthly": [100, 100, 100, 100, null, null, null, null, null, null, null, null],
          "average": 100,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 4,
          "name": "Number of Trips",
          "alias": "trips",
          "unit": "nos",
          "monthly": [25, 32, 42, 25, null, null, null, null, null, null, null, null],
          "average": 33.5,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 5,
          "name": "Quantity Shipped",
          "alias": "quantityShipped",
          "unit": "nos",
          "monthly": [124746, 110312, 226738, 91743, null, null, null, null, null, null, null, null],
          "average": 138884.75,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 6,
          "name": "No. of Parts per Trip",
          "alias": "partsPerTrip",
          "unit": "nos",
          "monthly": [4989.84, 3447.25, 5398.52, 3669.72, null, null, null, null, null, null, null, null],
          "average": 4376.83,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 7,
          "name": "Vehicle Turnaround Time",
          "alias": "vehicleTAT",
          "unit": "hrs",
          "monthly": [5, 4, 4, 4, null, null, null, null, null, null, null, null],
          "average": 4.33,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 8,
          "name": "Machine Downtime",
          "alias": "machineDowntimeHrs",
          "unit": "hrs",
          "monthly": [4, 3, 3.87, 0, null, null, null, null, null, null, null, null],
          "average": 2.72,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 9,
          "name": "Number of Machine Breakdowns",
          "alias": "machineBreakdowns",
          "unit": "nos",
          "monthly": [0.5, 0.2, 0.16, 0, null, null, null, null, null, null, null, null],
          "average": 0.215,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 10,
          "name": "On-Time Delivery Rate",
          "alias": "onTimeDeliveryRate",
          "unit": "%",
          "monthly": [95, 100, 100, 90, null, null, null, null, null, null, null, null],
          "average": 96.25,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 11,
          "name": "CAPA Closure Time",
          "alias": "capaClosureDays",
          "unit": "days",
          "monthly": [5, 3, 4, 2, null, null, null, null, null, null, null, null],
          "average": 3.5,
          "responsible": "Mr. Rohan",
          "remarks": ""
        }},
        {{
          "id": 12,
          "name": "Cost Per Part",
          "alias": "costPerPart",
          "unit": "INR",
          "monthly": [1.2, 1.1, 1.15, 1.05, null, null, null, null, null, null, null, null],
          "average": 1.125,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 13,
          "name": "Supplier Contribution % to Total Supply",
          "alias": "supplierContributionPercent",
          "unit": "%",
          "monthly": [12.5, 11.7, 13.2, 10.8, null, null, null, null, null, null, null, null],
          "average": 12.05,
          "responsible": "System Calculated",
          "remarks": ""
        }}
      ]
    }},
    {{
      "name": "XYZ Components",
      "location": "Pune",
      "buyer": "Ms. Swati Deshmukh",
      "spoc": "Mr. Prasad Kale",
      "kpis": [
        {{
          "id": 1,
          "name": "Safety - Accident Data",
          "alias": "accidents",
          "unit": "nos",
          "monthly": [0, 0, 0, 0, null, null, null, null, null, null, null, null],
          "average": 0,
          "responsible": "Mr. Rohan",
          "remarks": ""
        }},
        {{
          "id": 2,
          "name": "Production Loss due to Material Shortage",
          "alias": "productionLossHrs",
          "unit": "hrs",
          "monthly": [0, 0, 0, 0, null, null, null, null, null, null, null, null],
          "average": 0,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 3,
          "name": "OK Delivery Cycles (ACMA Standard)",
          "alias": "okDeliveryPercent",
          "unit": "%",
          "monthly": [100, 100, 100, 100, null, null, null, null, null, null, null, null],
          "average": 100,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 4,
          "name": "Number of Trips",
          "alias": "trips",
          "unit": "nos",
          "monthly": [25, 32, 42, 25, null, null, null, null, null, null, null, null],
          "average": 33.5,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 5,
          "name": "Quantity Shipped",
          "alias": "quantityShipped",
          "unit": "nos",
          "monthly": [124746, 110312, 226738, 91743, null, null, null, null, null, null, null, null],
          "average": 138884.75,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 6,
          "name": "No. of Parts per Trip",
          "alias": "partsPerTrip",
          "unit": "nos",
          "monthly": [4989.84, 3447.25, 5398.52, 3669.72, null, null, null, null, null, null, null, null],
          "average": 4376.83,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 7,
          "name": "Vehicle Turnaround Time",
          "alias": "vehicleTAT",
          "unit": "hrs",
          "monthly": [5, 4, 4, 4, null, null, null, null, null, null, null, null],
          "average": 4.33,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 8,
          "name": "Machine Downtime",
          "alias": "machineDowntimeHrs",
          "unit": "hrs",
          "monthly": [4, 3, 3.87, 0, null, null, null, null, null, null, null, null],
          "average": 2.72,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 9,
          "name": "Number of Machine Breakdowns",
          "alias": "machineBreakdowns",
          "unit": "nos",
          "monthly": [0.5, 0.2, 0.16, 0, null, null, null, null, null, null, null, null],
          "average": 0.215,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 10,
          "name": "On-Time Delivery Rate",
          "alias": "onTimeDeliveryRate",
          "unit": "%",
          "monthly": [95, 100, 100, 90, null, null, null, null, null, null, null, null],
          "average": 96.25,
          "responsible": "Mr. Hemant",
          "remarks": ""
        }},
        {{
          "id": 11,
          "name": "CAPA Closure Time",
          "alias": "capaClosureDays",
          "unit": "days",
          "monthly": [5, 3, 4, 2, null, null, null, null, null, null, null, null],
          "average": 3.5,
          "responsible": "Mr. Rohan",
          "remarks": ""
        }},
        {{
          "id": 12,
          "name": "Cost Per Part",
          "alias": "costPerPart",
          "unit": "INR",
          "monthly": [1.2, 1.1, 1.15, 1.05, null, null, null, null, null, null, null, null],
          "average": 1.125,
          "responsible": "Mr. Kirti",
          "remarks": ""
        }},
        {{
          "id": 13,
          "name": "Supplier Contribution % to Total Supply",
          "alias": "supplierContributionPercent",
          "unit": "%",
          "monthly": [12.5, 11.7, 13.2, 10.8, null, null, null, null, null, null, null, null],
          "average": 12.05,
          "responsible": "System Calculated",
          "remarks": ""
        }}
      ]
    }}
  ]
}}
Fill in monthly data with null for missing months. Ensure valid JSON format."""

    user_prompt = f"""
Raw KPI data:
{json.dumps(raw_supplier_kpi_data, indent=2)}
Now convert this to the final dashboard structure for all suppliers.
"""

    try:
        response = client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[
                {"role": "system", "content": system_prompt.strip()},
                {"role": "user", "content": user_prompt.strip()}
            ],
            temperature=0.2,
            max_tokens=4096
        )
        reply = response.choices[0].message.content
        return json.loads(reply)
    except Exception as e:
        print("❌ Error during LLM generation:", e)
        return {}

def dashboard_data(file_path, output_path="results/dashboard.json"):
    """
    Extract insights data from Excel file and save to JSON format.
    Args:
        file_path: Path to the Excel file
        output_path: Path where the JSON output should be saved (default: results/dashboard.json)
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        if len(sheet_names) > 1:
            sheet_names = sheet_names[1:]

        insights = defaultdict(dict)  # {metric: {company: {month: value}}}

        print(f"📊 Processing {len(sheet_names)} sheets from Excel file...")

        for sheet_index, sheet_name in enumerate(sheet_names):
            try:
                sheet = wb[sheet_name]
                company = sheet_name.strip()

                print(f"🔄 Processing sheet: {company}")

                header_row_num = None
                header = None
                month_indices = []

                for row_num in range(1, 11):
                    try:
                        row_cells = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                        if not row_cells:
                            continue

                        row_strings = [str(cell).strip().title() if cell else "" for cell in row_cells]
                        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

                        found_months = [(i, cell_text) for i, cell_text in enumerate(row_strings) if cell_text in months]

                        if found_months:
                            header_row_num = row_num
                            header = row_strings
                            month_indices = found_months
                            break

                    except Exception:
                        continue

                if not header_row_num or not month_indices:
                    print(f"⚠️ No month headers found in sheet: {company}")
                    continue

                data_rows_processed = 0
                for row_num, row in enumerate(sheet.iter_rows(min_row=header_row_num + 1, values_only=True), header_row_num + 1):
                    if not row or len(row) < 2:
                        continue

                    metric_cell = row[1] if len(row) > 1 else None
                    if not metric_cell:
                        continue

                    metric = str(metric_cell).strip()
                    if not metric or metric.lower() in ['', 'none', 'null']:
                        continue

                    values = {}
                    for col_idx, month in month_indices:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val is not None:
                                try:
                                    if isinstance(val, (int, float)) and val != 0:
                                        values[month] = float(val)
                                    elif isinstance(val, str):
                                        clean_val = val.replace(',', '').replace('$', '').replace('%', '').strip()
                                        if clean_val and clean_val not in ['-', '0', '0.0']:
                                            float_val = float(clean_val)
                                            if float_val != 0:
                                                values[month] = float_val
                                except (ValueError, TypeError):
                                    continue

                    if values:
                        insights[metric][company] = values
                        data_rows_processed += 1

                print(f"✅ Processed {data_rows_processed} data rows for {company}")

            except Exception as e:
                print(f"❌ Error processing sheet '{sheet_name}': {e}")
                continue

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        insights_dict = dict(insights)
        for metric in insights_dict:
            insights_dict[metric] = dict(insights_dict[metric])

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(insights_dict, f, indent=2, ensure_ascii=False)

        print(f"✅ Raw extracted dashboard saved to {output_path}")

        # Now run LLM generation
        structured = generate_dashboard_json_with_llm(insights_dict)
        structured_path = Path("results/structured_dashboard.json")
        with open(structured_path, "w", encoding="utf-8") as f:
            json.dump(structured, f, indent=2, ensure_ascii=False)

        print(f"🧠 Structured dashboard saved to {structured_path}")
        return structured

    except Exception as e:
        print(f"❌ Error processing Excel file: {e}")
        raise

